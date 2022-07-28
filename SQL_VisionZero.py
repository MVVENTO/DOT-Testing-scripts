#Author: Melissa A Vento
#Date: 06/29/2022
# Description: python script to automate an excel spreedsheet
#              and dynamically update a new Column for each Month coming up
#              Fully Automate this process using Database

#Install Libraires to create an excel spreedsheet
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Alignment

import os
#import arcpy
#from arcpy import env
#arcpy.env.overwriteOutput = True
from datetime import datetime
import pyodbc



#connect to sql database
connection_string = ("Driver={SQL Server};""Server=dotgissql01;""Database=VisionZeroView;""UID=gisuser;""PWD=GISUSER;")
connection = pyodbc.connect(connection_string)
print("connected to sql database")



data = {

    # New speed humps installed
    # Database:
    # Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Safety_Interventions\VisionZeroView.GISADMIN.VZV_Speed_Humps
    "5": {
    "Metric Name":"New speed humps installed",
	"January 2022 (spatial)":0,
	"February 2022 (spatial)":"N/A",
	"March 2022 (spatial)":0,
	"April 2022 (spatial)":0,
	"May 2022 (spatial)":0,
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":0, # sum of Columns C2, D2, E2, F2, G2,
    "Overall Target":250,
    "Primary Contacts":"Ann Marie Doherty/ Alicia Posner/ Seth Hostetter",
    "Data Contacts":"Seth Hostetter/Arthur Getman"
    },

    # Safety projects (SIPs) completed _ VZ priority geographies only
    # Database:
    # N/A
    "7": {
    "Metric Name":"Safety projects (SIPs) completed _ VZ priority geographies only",
	"January 2022 (spatial)":0,
	"February 2022 (spatial)":"N/A",
	"March 2022 (spatial)":"N/A",
	"April 2022 (spatial)":"N/A",
	"May 2022 (spatial)":"N/A",
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":"N/A",  
    "Overall Target":50,
    "Primary Contacts":"Jeff Malamy/ Sean Quinn",
    "Data Contacts":"Seth Hostetter/Arthur Getman"
    },


    #Safety projects (SIPs) completed _ all projects
    # Databases:
    # Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Safety_Interventions\VisionZeroView.GISADMIN.VZV_SIP_Corridors
    # Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Safety_Interventions\VisionZeroView.GISADMIN.VZV_SIP_Intersections
    "8": {
    "Metric Name":"Safety projects (SIPs) completed _ all projects ",
	"January 2022 (spatial)":3, # sum of SIP_ Corridors + SIP_Intersections 
	"February 2022 (spatial)":1, # sum of SIP_ Corridors + SIP_Intersections
	"March 2022 (spatial)":0, # sum of SIP_ Corridors + SIP_Intersections
	"April 2022 (spatial)":2, # sum of SIP_ Corridors + SIP_Intersections
	"May 2022 (spatial)":0,
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":6, # sum of Columns C4, D4, E4, F4, G4,
    "Overall Target":250,
    "Primary Contacts":"Jeff Malamy/ Sean Quinn",
    "Data Contacts":"Seth Hostetter/Arthur Getman"
    },

    #Enhanced crossings completed
    # Database:
    #Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Safety_Interventions\VisionZeroView.GISADMIN.VZV_Enhanced_Crossings
    "9": {
    "Metric Name":"Enhanced crossings completed",
	"January 2022 (spatial)":0,  #sum of Enhanced_Crossings
	"February 2022 (spatial)":0,  #sum of Enhanced_Crossings
	"March 2022 (spatial)":0,  #sum of Enhanced_Crossings
	"April 2022 (spatial)":0,  #sum of Enhanced_Crossings
	"May 2022 (spatial)":0,  #sum of Enhanced_Crossings
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":0, # sum of Columns C5, D5, E5, F5, G5,
    "Overall Target":25,
    "Primary Contacts":"Heidi Wolf/ Terra Ishee/ Julio Palleiro",
    "Data Contacts":"Julio Palleiro"
    },

    #Turn traffic calming treatments installed
    # Database:
    #Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Safety_Interventions\VisionZeroView.GISADMIN.VZV_Turn_Traffic_Calming
    #completion

    "10": {
    "Metric Name":"Turn traffic calming treatments installed",
	"January 2022 (spatial)":1, # sum of  Turn_Traffic_Calming
	"February 2022 (spatial)":0, # sum of Turn_Traffic_Calming
	"March 2022 (spatial)":0, # sum of Turn_Traffic_Calming
	"April 2022 (spatial)":26, # sum of Turn_Traffic_Calming
	"May 2022 (spatial)":0,
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":27, # sum of Columns C6, D6, E6, F6, G6,
    "Overall Target":100,
    "Primary Contacts":"Arban Vigani/ Rob Viola",
    "Data Contacts":"Seth Hostetter/Arthur Getman"
    },

    # New traffic signals installed
    # Database:
    # Database Connections\GISGRID on DOTGISSQL01.sde\GISGRID.GISADMIN.TRAFFIC_PLANNING\GISGRID.GISADMIN.Signal_Controller
    "13": {
    "Metric Name":"New traffic signals installed",
	"January 2022 (spatial)": 0, # sum of  Signal_Controller
	"February 2022 (spatial)":0, # sum of Signal_Controller
	"March 2022 (spatial)":0, # sum of Signal_Controller
	"April 2022 (spatial)":0, # sum of Signal_Controller
	"May 2022 (spatial)":0,
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":0, # sum of Columns C7, D7, E7, F7, G7,
    "Overall Target":70,
    "Primary Contacts":"Ernie Athanailos/ Dyesha Mitchell",
    "Data Contacts":"Cindi Ochs/Milton Nyazema"
    },

    # Leading pedestrian intervals installed
    # Database:
    # Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Safety_Interventions\VisionZeroView.GISADMIN.VZV_Leading_Pedestrian_Intervals
    "14": {
    "Metric Name":"Leading pedestrian intervals installed",
	"January 2022 (spatial)":6, # sum of Leading_Pedestrian_Intervals 
	"February 2022 (spatial)":81, # sum of Leading_Pedestrian_Intervals
	"March 2022 (spatial)":75, # sum of Leading_Pedestrian_Intervals
	"April 2022 (spatial)":123, # sum of Leading_Pedestrian_Intervals
	"May 2022 (spatial)":0,
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":285, # sum of Columns C8, D8, E8, F8, G8,
    "Overall Target":700,
    "Primary Contacts":"Ernie Athanailos/ Dan Nguyen",
    "Data Contacts":"Jason Fitzsimmons"
    },

    # Schools with targeted educational programs and outreach completed
    # Database:
    # Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Outreach\VisionZeroView.GISADMIN.VZV_Outreach_Schools
    "17": {
    "Metric Name":"Schools with targeted educational programs and outreach completed",
	"January 2022 (spatial)":39, # sum of Outreach_Schools
	"February 2022 (spatial)":46, # sum of Outreach_Schools
	"March 2022 (spatial)":60, # sum of Outreach_Schools
	"April 2022 (spatial)":235, # sum of Outreach_Schools
	"May 2022 (spatial)":0,
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":180, # sum of Columns C9, D9, E9, F9, G9,
    "Overall Target":500,
    "Primary Contacts":"Lillie Mitchell/ Marianne Groomes",
    "Data Contacts":"Jason Fitzsimmons"
    },

    # Senior centers partnered with at priority/high-crash neighborhoods
    # Database:
    # Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Outreach\VisionZeroView.GISADMIN.VZV_Senior_Centers
    "18": {
    "Metric Name": "Senior centers partnered with at priority/high-crash neighborhoods",
	"January 2022 (spatial)":9, # sum of Senior_Centers
	"February 2022 (spatial)":22, # sum of Senior_Centers
	"March 2022 (spatial)":17, # sum of Senior_Centers
	"April 2022 (spatial)":35, # sum of Senior_Centers
	"May 2022 (spatial)":0,
        # dynamically add months to python script
	"YTD Total 2022 (spatial)":83, # sum of Columns C10, D10, E10, F10, G10,
    "Overall Target":100,
    "Primary Contacts":"Lillie Mitchell/ Marianne Groomes",
    "Data Contacts":"Jason Fitzsimmons"
    },

    # Events conducted for hands-on safety demonstrations
    # Database:
    # Database Connections\VisionZeroView on DOTGISSQL01.sde\VisionZeroView.GISADMIN.Outreach\VisionZeroView.GISADMIN.VZV_Hands_On_Safety_Demos
    "19": {
    "Metric Name": "Events conducted for hands-on safety demonstrations",
    "January 2022 (spatial)":4, # sum of Hands_On_Safety_Demos
    "February 2022 (spatial)":7, # sum of Hands_On_Safety_Demos
    "March 2022 (spatial)":7, # sum of Hands_On_Safety_Demos
    "April 2022 (spatial)":27, # sum of Hands_On_Safety_Demos
    "May 2022 (spatial)":0,
        # dynamically add months to python script
     "YTD Total 2022 (spatial)":45, # sum of Columns C11, D11, E11, F11, G11,
    "Overall Target": 75,
    "Primary Contacts":"Lillie Mitchell/ Marianne Groomes",
    "Data Contacts":"Jason Fitzsimmons"
    }
}

#Titles = {"20": {"JUNE 2022 (spatial)": 0}, "21": {"JULY 2022 (spatial)": 0}}

#Updates dictionary
#data.update(Titles)



# Load new workbook titled "Vision Zero Metrics"
wb = Workbook()
ws = wb.active
ws.title = "Vision Zero Metrics"



#Display headings
headings = ['No.'] + list(data['5'].keys())
ws.append(headings)


#prints Values into rows
for Metric_name in data:
    metrics = list(data[Metric_name].values())
    ws.append([Metric_name]+ metrics)



#### adding new Months example 1 ####
#new_information = {"JUNE 2022 (spatial)": 0}
#data.update(new_information)
#print(data)

#### adding new Months example 2 ####
#merging and updating dictionaries with | and |=
#JUNE 2022 (spatial) = {"number value"}
#JULY 2022 (spatial) = {"number value"}
#JUNE 2022 (spatial) | JULY 2022 (spatial)


'''
# add new column for June
ws.insert_cols(8)


# add new column for July
ws.insert_cols(9)

'''



############ FORMAT #############


#Makes Titles black and bold 
#             cells filled color 0099CCFF
#             Font size 12 and name Leelawadee 

for col in range(1, 14):
	ws[get_column_letter(col) + '1'].fill = PatternFill(start_color='0099CCFF', fill_type='solid')
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="00000000", size='12', name='Leelawadee')
	ws[get_column_letter(col) + '1'].alignment = Alignment(wrap_text=True,  vertical='center')


#YTD row font color
#7030A0
for row in range(9,10):
        ws[get_column_letter(row) + '1'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')



#2 row
for col in range(1,3):
        ws[get_column_letter(col) + '2'].font = Font(color="00000000", size='12', name='Leelawadee')

#Makes Values red and bold 
for col in range(3,9):
        ws[get_column_letter(col) + '2'].font = Font(color="00FF0000",bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '2'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '2'].font = Font(color="00000000", size='12', name='Leelawadee')


#3 row
for col in range(1,3):
        ws[get_column_letter(col) + '3'].font = Font(color="00000000", size='12', name='Leelawadee')
for col in range(3,9):
        ws[get_column_letter(col) + '3'].font = Font(color="00FF0000",bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '3'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '3'].font = Font(color="00000000", size='12', name='Leelawadee')


#4 row
for col in range(1,3):
        ws[get_column_letter(col) + '4'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '4'].font = Font(color="00FF0000", bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '4'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '4'].font = Font(color="00000000", size='12', name='Leelawadee')

#5 row
for col in range(1,3):
        ws[get_column_letter(col) + '5'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '5'].font = Font(color="00FF0000",bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '5'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '5'].font = Font(color="00000000", size='12', name='Leelawadee')


#6 row
for col in range(1,3):
        ws[get_column_letter(col) + '6'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '6'].font = Font(color="00FF0000", bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '6'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '6'].font = Font(color="00000000", size='12', name='Leelawadee')


#7 row
for col in range(1,3):
        ws[get_column_letter(col) + '7'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '7'].font = Font(color="00FF0000",bold=True,  size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '7'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '7'].font = Font(color="00000000", size='12', name='Leelawadee')


#8 row
for col in range(1,3):
        ws[get_column_letter(col) + '8'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '8'].font = Font(color="00FF0000", bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '8'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '8'].font = Font(color="00000000", size='12', name='Leelawadee')

#9 row 
for col in range(1,3):
        ws[get_column_letter(col) + '9'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '9'].font = Font(color="00FF0000",bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '9'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '9'].font = Font(color="00000000", size='12', name='Leelawadee')


#10 row
for col in range(1,3):
        ws[get_column_letter(col) + '10'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '10'].font = Font(color="00FF0000", bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '10'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(10,13):
        ws[get_column_letter(col) + '10'].font = Font(color="00000000", size='12', name='Leelawadee')


#11 row
for col in range(1,3):
        ws[get_column_letter(col) + '11'].font = Font(color="00000000", size='12', name='Leelawadee')

for col in range(3,9):
        ws[get_column_letter(col) + '11'].font = Font(color="00FF0000", bold=True, size='12', name='Leelawadee')

# YTD Bold purple
for row in range(9,10):
        ws[get_column_letter(row) + '11'].font = Font(color="7030A0",bold=True, size='12', name='Leelawadee')

for col in range(11,13):
        ws[get_column_letter(col) + '11'].font = Font(color="00000000", size='12', name='Leelawadee')





    



#save Spreedsheet
wb.save("Vision Zero Metrics_test.xlsx")
